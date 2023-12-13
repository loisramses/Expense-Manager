import matplotlib.figure as fig
import openpyxl as xlrw
import tkinter as tk
import pandas as pd
import numpy as np
import mplcursors
import os
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
from openpyxl.styles import Font
from tkinter import messagebox
from datetime import datetime
from datetime import date
from tkinter import ttk

class Manager:
    def __init__(self):
        self.log = []
        self.root = tk.Tk()
        self.root.title("Expense Manager")
        self.root.geometry("1200x700")

        self.folder_path = './expenses/'
        self.performing_operation = False

        self.main_font = ('Arial', 13)
        self.button_font = ('Arial', 9)
        self.cell_font = Font(name='Calibri', size=11, bold=True)
        self.ammount_format_str = '#,##0.00 €; [Red]-#,##0.00 €'
        self.date_format_str = 'DD/MM/YYYY'
        self.categories_list = ['Education', 'Food', 'Health', 'Income', 'Investments', 'Leisure', 'Shopping', 'Subscription', 'Transportation', 'Travel', 'Other']
        # self.categories_list = ['Alimentação', 'Compras', 'Educação', 'Investimentos', 'Lazer', 'Saúde', 'Transportes', 'Viagens', 'Outro']
        mplcursors.cursor(hover=True)

        self.setup_menubar()
        self.setup_excelExplorerFrame()
        self.setup_statisticsFrame()

        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

    def setup_menubar(self):
        self.menubar = tk.Menu(self.root)
        
        self.file_menu = tk.Menu(self.menubar, tearoff=0)
        self.file_menu.add_command(label="Create Year", command=self.add_year)
        self.file_menu.add_command(label="Create Month", command=self.add_month)
        self.file_menu.add_command(label="Edit Month", command=self.edit_month)
        self.file_menu.add_command(label="Delete Year", command=self.delete_year)
        self.file_menu.add_command(label="Delete Month", command=self.delete_month)
        self.file_menu.add_command(label="Close", command=self.on_closing)
        self.menubar.add_cascade(menu=self.file_menu, label="File")

        self.action_menu = tk.Menu(self.menubar, tearoff=0)
        self.action_menu.add_command(label="Add Expense", command=lambda: self.add_dataToExcel('expense'))
        self.action_menu.add_command(label="Add Revenue", command=lambda: self.add_dataToExcel('revenue'))
        self.action_menu.add_command(label="Edit Selection", command=self.edit_data)
        self.action_menu.add_command(label="Delete Selection", command=self.delete_selection)
        self.menubar.add_cascade(menu=self.action_menu, label="Action")

        self.root.config(menu=self.menubar)

    def create_workbook(self, info):
        self.add_log(f'Creating workbook with name "{info["year_name"]}"')

        new_file_name = info['year_name'] + '.xlsx'
        info['year_name'] = new_file_name

        # create a new workbook, a new sheet and save it
        self.current_workbook = xlrw.Workbook()
        self.current_sheet = self.current_workbook.active
        self.create_worksheet(info)

        # update workbook list
        self.books_names = os.listdir('./expenses')
        self.year_boxlist.config(values=self.books_names)
        self.year_boxlist.current(self.books_names.index(new_file_name))

        self.update_sheets()

    def add_year(self):
        self.perform_op()
        year_frame = addYear.AddYear(self)
        year_frame.run()

    def delete_year(self):
        year = self.year_boxlist.get()
        if messagebox.askyesno("Confirm", f"Are you sure you want to delete {year}?"):
            self.add_log(f'Deleting {year}')
            os.remove(self.folder_path + year)

            # update workbook list
            self.books_names = os.listdir('./expenses')
            self.year_boxlist.config(values=self.books_names)
            self.year_boxlist.current(0)

            self.update_sheets()

    def create_worksheet(self, info):
        self.add_log(f'Creating worksheet "{info["month_name"]}" with initial date "{info["init_date"]}" and initial amount "{info["init_ammount"]}" on current workbook "{info["year_name"]}"')

        self.current_sheet.title = info['month_name']

        # insert col tags
        self.current_sheet['A1'] = 'DATE'
        self.current_sheet['A1'].font = self.cell_font
        self.current_sheet['B1'] = 'AMOUNT'
        self.current_sheet['B1'].font = self.cell_font
        self.current_sheet['C1'] = 'PURPOSE'
        self.current_sheet['C1'].font = self.cell_font
        self.current_sheet['D1'] = 'DESCRIPTION'
        self.current_sheet['D1'].font = self.cell_font
        self.current_sheet['E1'] = 'CATEGORY'
        self.current_sheet['E1'].font = self.cell_font

        # insert dates
        self.current_sheet['G1'] = 'Data início:'
        self.current_sheet['G1'].font = self.cell_font
        self.current_sheet['H1'] = info['init_date']
        self.current_sheet['G2'] = 'Data fim:'
        self.current_sheet['G2'].font = self.cell_font

        # insert ammounts
        self.current_sheet['G4'] = 'Montante inicial:'
        self.current_sheet['G4'].font = self.cell_font
        self.current_sheet['H4'] = info['init_ammount']
        self.current_sheet['H4'].number_format = self.ammount_format_str
        self.current_sheet['G5'] = 'Montante final:'
        self.current_sheet['G5'].font = self.cell_font
        self.current_sheet['H5'] = '=(H4-H7)'
        self.current_sheet['H5'].number_format = self.ammount_format_str

        # total spent ammount
        self.current_sheet['G7'] = 'Total gasto:'
        self.current_sheet['G7'].font = self.cell_font
        self.current_sheet['H7'] = '=SUM(B:B)'
        self.current_sheet['H7'].number_format = self.ammount_format_str

        self.current_workbook.save(self.folder_path + info['year_name'])

    def add_month(self):
        self.perform_op()
        month_frame = addMonth.AddMonth(self)
        month_frame.run()

    def edit_current_sheet(self, info):
        self.add_log(f'Editing worksheet "{self.month_boxlist.get()}":')
        self.add_log(f'previous info - start_date: {self.current_sheet_start_date}, end_date: {self.current_sheet_end_date}, initial_amount: {self.current_sheet_initial_ammount}')

        # set the updated data for the current sheet
        self.current_sheet_start_date = info['start_date']
        self.current_sheet_end_date = info['end_date']
        self.current_sheet_initial_ammount = info['initial_ammount']
        self.add_log(f'updated info - start_date: {self.current_sheet_start_date}, end_date: {self.current_sheet_end_date}, initial_amount: {self.current_sheet_initial_ammount}')
        self.save_workbook()

    def edit_month(self):
        self.perform_op()
        editMonth_frame = editMonth.EditMonth(self)
        editMonth_frame.run()

    def delete_month(self):
        month = self.month_boxlist.get()
        if messagebox.askyesno("Confirm", f"Are you sure you want to delete {month}?"):
            self.add_log(f'Deleting {month}')
            self.current_workbook.remove(self.current_sheet)
            self.save_workbook()
            self.update_sheets()

    def setup_excelExplorerFrame(self):
        self.books_names = os.listdir(self.folder_path)

        self.excel_explorerFrame = tk.Frame(self.root)

        self.year_boxlist = ttk.Combobox(self.excel_explorerFrame, value=self.books_names, state='readonly')
        self.year_boxlist.current(0)
        self.update_sheets()
        self.year_boxlist.grid(row=0, column=0, padx=5, pady=5, sticky=tk.EW)
        self.year_boxlist.bind("<<ComboboxSelected>>", self.update_sheets)

        self.month_boxlist = ttk.Combobox(self.excel_explorerFrame, value=self.sheet_names, state='readonly')
        self.month_boxlist.current(0)
        self.month_boxlist.grid(row=0, column=1, padx=5, pady=5, sticky=tk.EW)
        self.month_boxlist.bind("<<ComboboxSelected>>", self.update_expenses)

        # self.headings = ["DATE", "AMMOUNT", "PURPOSE", "DESCRIPTION", "CATEGORY"]

        self.update_expenses()
        self.expense_treeList = ttk.Treeview(self.excel_explorerFrame, columns=self.headings, show='headings')
        for i in range(0, 5): self.expense_treeList.column(i, anchor=tk.CENTER)
        for col in self.headings: self.expense_treeList.heading(col, text=col)
        self.populate_expensesTree()
        self.expense_treeList.grid(row=1, column=0, rowspan=4, columnspan=2, padx=5, pady=5, sticky=tk.NSEW)

        # add expense button
        self.add_button = tk.Button(self.excel_explorerFrame, text="Add Expense", bg='firebrick2', font=('Arial', 9), width=13, command=lambda: self.add_dataToExcel('expense'))
        self.add_button.grid(row=1, column=2, padx=5, pady=5, sticky=tk.EW)

        # add revenue button
        self.add_button = tk.Button(self.excel_explorerFrame, text="Add Revenue", bg='green2', font=('Arial', 9), width=13, command=lambda: self.add_dataToExcel('revenue'))
        self.add_button.grid(row=2, column=2, padx=5, pady=5, sticky=tk.EW)

        # edit button
        self.edit_button = tk.Button(self.excel_explorerFrame, text="Edit", bg='light blue', font=('Arial', 9), command=self.edit_data)
        self.edit_button.grid(row=3, column=2, padx=5, pady=5, sticky=tk.EW)

        # delete button
        self.delete_button = tk.Button(self.excel_explorerFrame, text="Delete Selection", bg='red2', font=('Arial', 9), command=self.delete_selection)
        self.delete_button.grid(row=4, column=2, padx=5, pady=5, sticky=tk.EW)

        self.excel_explorerFrame.pack()

    def update_sheets(self, event=None, index=0):
        # close the previous book
        if hasattr(self, 'workbook'):
            self.current_workbook.close()

        self.current_workbook = xlrw.load_workbook(self.folder_path + self.year_boxlist.get())
        self.sheet_names = self.current_workbook.sheetnames
        if hasattr(self, 'month_boxlist'):
            self.month_boxlist.config(value=self.sheet_names)
            self.month_boxlist.current(index)
            self.update_expenses()

    def update_expenses(self, event=None):
        self.row_nmb = 0
        if hasattr(self, 'current_sheet'):
            self.save_workbook() # save information when changing worksheets
        
        self.current_sheet = self.get_current_sheet()
        self.current_sheet_data = self.get_current_month_data(self.current_sheet)

        if hasattr(self, 'category_stat_boxlist'):
            self.update_all_stats()
        self.populate_expensesTree()

    def populate_expensesTree(self):
        if hasattr(self, 'expense_treeList'):
            self.expense_treeList.delete(*self.expense_treeList.get_children())
            for row in self.current_sheet_data:
                self.expense_treeList.insert("", 'end', values=row)

    def add_row_to_current_sheet(self, row):
        # add row to excel
        self.current_sheet.cell(row=self.row_nmb+1, column=1, value=row[0])
        try:
            self.current_sheet.cell(row=self.row_nmb+1, column=2, value=float(row[1])).number_format = self.ammount_format_str
        except:
            messagebox.showwarning("Invalid number format", "The value must be seperated by \".\"!")
            return
        self.current_sheet.cell(row=self.row_nmb+1, column=3, value=str(row[2]))
        self.current_sheet.cell(row=self.row_nmb+1, column=4, value=str(row[3]))
        self.current_sheet.cell(row=self.row_nmb+1, column=5, value=str(row[4]))

        self.current_sheet_data.append(row)
        self.save_workbook()
        self.row_nmb += 1
        self.populate_expensesTree()

    def add_dataToExcel(self, type_of_op):
        self.perform_op()
        add_frame = addFrame.AddData(self, type_of_op)
        add_frame.run()
        if hasattr(add_frame, 'new_row'):
            self.add_log(f'Added "{type_of_op}" with data {add_frame.new_row}')

    def edit_row_on_current_sheet(self, row):
        date = row[0]
        ammount = row[1]
        purpose = row[2]
        description = row[3]
        category = row[4]

        # update expense tree list
        self.expense_treeList.set(item=self.selection[0], column=0, value=date)
        self.expense_treeList.set(item=self.selection[0], column=1, value=ammount)
        self.expense_treeList.set(item=self.selection[0], column=2, value=purpose)
        self.expense_treeList.set(item=self.selection[0], column=3, value=description)
        self.expense_treeList.set(item=self.selection[0], column=4, value=category)

        # update excel
        target_index = self.expense_treeList.index(self.selection[0])+2
        self.current_sheet.cell(row=target_index, column=1, value=date)
        try:
            self.current_sheet.cell(row=target_index, column=2, value=ammount).number_format = self.ammount_format_str
        except:
            messagebox.showwarning("Invalid number format", "The value must be seperated by \".\"!")
            return
        self.current_sheet.cell(row=target_index, column=3, value=purpose)
        self.current_sheet.cell(row=target_index, column=4, value=description)
        self.current_sheet.cell(row=target_index, column=5, value=category)

        self.save_workbook()

    def edit_data(self):
        self.selection = self.expense_treeList.selection()
        selection_size = len(self.selection)
        if selection_size == 0:
            messagebox.showwarning("Empty selection", "There are no items selected, cannot perform operation!")
            return
        if selection_size > 1:
            messagebox.showwarning("Too many items selected", "There are too many items selected, cannot perform operation! Edit can only be performed on one item.")
            return
        self.perform_op()
        edit_frame = editFrame.EditExpense(self)
        edit_frame.run()
        if hasattr(edit_frame, 'item_values') and hasattr(edit_frame, 'new_row'):
            self.add_log(f'Editing {edit_frame.item_values} to {edit_frame.new_row}')

    def delete_selection(self):
        selection = self.expense_treeList.selection()
        selection_size = len(selection)
        if selection_size <= 0:
            messagebox.showwarning("Empty selection", "There are no items selected, cannot perform operation!")
            return
        if messagebox.askyesno("Confirm deletion", "Are you sure you want to delete all the selected items? This will permanently delete the selected items!"):
            for row in selection:
                self.add_log(f'Deleting {self.expense_treeList.item(row)}')
                self.current_sheet.delete_rows(self.expense_treeList.index(row)+1, 1) # +1 bc of header and index starts at 0
                self.expense_treeList.delete(row)
                self.row_nmb -= 1

            self.update_expenses()
            self.populate_expensesTree()
            self.save_workbook()

    def get_current_sheet(self):
        sheet = self.current_workbook[self.month_boxlist.get()]
        self.current_sheet_start_date = sheet['H1'].value
        self.current_sheet_end_date = sheet['H2'].value
        self.current_sheet_initial_ammount = sheet['H4'].value
        return sheet

    def save_workbook(self):
        self.add_log(f'Saving workbook {self.year_boxlist.get()}')

        # clear old data
        self.current_sheet.delete_cols(7, 2)

        # insert dates
        self.current_sheet['G1'] = 'Data início:'
        self.current_sheet['G1'].font = self.cell_font
        self.current_sheet['H1'] = self.current_sheet_start_date
        self.current_sheet['H1'].number_format = self.date_format_str
        self.current_sheet['G2'] = 'Data fim:'
        self.current_sheet['G2'].font = self.cell_font
        self.current_sheet['H2'] = self.current_sheet_end_date
        self.current_sheet['H2'].number_format = self.date_format_str

        # insert ammounts
        self.current_sheet['G4'] = 'Montante inicial:'
        self.current_sheet['G4'].font = self.cell_font
        self.current_sheet['H4'] = self.current_sheet_initial_ammount
        self.current_sheet['H4'].number_format = self.ammount_format_str
        self.current_sheet['G5'] = 'Montante final:'
        self.current_sheet['G5'].font = self.cell_font
        self.current_sheet['H5'] = '=(H4-H7)'
        self.current_sheet['H5'].number_format = self.ammount_format_str

        # total spent ammount
        self.current_sheet['G7'] = 'Total gasto:'
        self.current_sheet['G7'].font = self.cell_font
        self.current_sheet['H7'] = '=SUM(B:B)'
        self.current_sheet['H7'].number_format = self.ammount_format_str

        self.current_workbook.save(self.folder_path + self.year_boxlist.get())

    def setup_statisticsFrame(self):
        self.statisticsFrame = tk.Frame(self.root)

        self.category_stat_boxlist = ttk.Combobox(self.statisticsFrame, value=['Current Month', 'Current Year'], state='readonly')
        self.category_stat_boxlist.current(0)
        self.category_stat_boxlist.grid(row=0, column=0,padx=5, pady=5, sticky=tk.EW)
        self.category_stat_boxlist.bind("<<ComboboxSelected>>", self.update_categoryStats)

        self.gain_spendings_stat_boxlist = ttk.Combobox(self.statisticsFrame, value=['Current Month', 'Current Year'], state='readonly')
        self.gain_spendings_stat_boxlist.current(0)
        self.gain_spendings_stat_boxlist.grid(row=0, column=1,padx=5, pady=5, sticky=tk.EW)
        self.gain_spendings_stat_boxlist.bind("<<ComboboxSelected>>", self.update_gain_spendingsStats)

        self.daily_spendings_stat_boxlist = ttk.Combobox(self.statisticsFrame, value=['Current Month', 'Current Year'], state='readonly')
        self.daily_spendings_stat_boxlist.current(0)
        self.daily_spendings_stat_boxlist.grid(row=0, column=2,padx=5, pady=5, sticky=tk.EW)
        self.daily_spendings_stat_boxlist.bind("<<ComboboxSelected>>", self.update_daily_spendingsStats)

        self.category_stat_fig = fig.Figure(figsize=(3.5, 3.5))
        self.category_stat_ax = self.category_stat_fig.add_subplot()
        self.category_stat_canvas = FigureCanvasTkAgg(self.category_stat_fig, self.statisticsFrame)
        self.category_stat_canvas.get_tk_widget().grid(row=1, column=0, padx=5, pady=5, sticky=tk.EW)

        self.gain_spendings_stat_fig = fig.Figure(figsize=(3.5, 3.5))
        self.gain_spendings_stat_ax = self.gain_spendings_stat_fig.add_subplot()
        self.gain_spendings_stat_canvas = FigureCanvasTkAgg(self.gain_spendings_stat_fig, self.statisticsFrame)
        self.gain_spendings_stat_canvas.get_tk_widget().grid(row=1, column=1, padx=5, pady=5, sticky=tk.EW)

        self.daily_spendings_stat_fig = fig.Figure(figsize=(3.5, 3.5))
        self.daily_spendings_stat_ax = self.daily_spendings_stat_fig.add_subplot()
        self.daily_spendings_stat_canvas = FigureCanvasTkAgg(self.daily_spendings_stat_fig, self.statisticsFrame)
        self.daily_spendings_stat_canvas.get_tk_widget().grid(row=1, column=2, padx=5, pady=5, sticky=tk.EW)

        self.update_all_stats()

        self.statisticsFrame.pack()

    def update_all_stats(self):
        self.update_categoryStats()
        self.update_gain_spendingsStats()
        self.update_daily_spendingsStats()

    def update_categoryStats(self, event=None):
        if (self.category_stat_boxlist.current() == 0): data = self.get_current_month_data(self.current_sheet)
        else: data = self.get_current_year_data(self.current_workbook)
        df = pd.DataFrame(data, columns=self.headings)
        df['AMMOUNT'] = df['AMMOUNT'].abs()
        cat_totals = df.groupby('CATEGORY').sum().round(2)
        self.category_stat_ax.clear()
        _, _, autotexts = self.category_stat_ax.pie(cat_totals['AMMOUNT'], labels=cat_totals.index, autopct='%.2f%%')
        for i, value in enumerate(cat_totals['AMMOUNT']):
            autotexts[i].set_text(str(value))
        self.category_stat_canvas.draw()

    def update_gain_spendingsStats(self, event=None):
        if (self.gain_spendings_stat_boxlist.current() == 0): data = self.get_current_month_data(self.current_sheet)
        else: data = self.get_current_year_data(self.current_workbook)
        df = pd.DataFrame(data, columns=self.headings)
        df['GROUP'] = df['CATEGORY'].apply(lambda x: 'Income' if x == 'Income' else 'Expenses')
        df['AMMOUNT'] = df['AMMOUNT'].abs()
        cat_totals = df.groupby('GROUP').sum().round(2)
        self.gain_spendings_stat_ax.clear()
        _, _, autotexts = self.gain_spendings_stat_ax.pie(cat_totals['AMMOUNT'], labels=cat_totals.index, autopct='%.2f%%')
        for i, value in enumerate(cat_totals['AMMOUNT']):
            autotexts[i].set_text(value)
        self.gain_spendings_stat_canvas.draw()

    def update_daily_spendingsStats(self, event=None):
        if (self.daily_spendings_stat_boxlist.current() == 0): data = self.get_current_month_data(self.current_sheet)
        else: data = self.get_current_year_data(self.current_workbook)
        df = pd.DataFrame(data, columns=self.headings)
        df['DATE'] = pd.to_datetime(df['DATE'], format='%d/%m/%Y')
        df['DATE'] = df['DATE'].dt.strftime('%d/%m')
        df['AMMOUNT'] = df['AMMOUNT'].abs()
        df['GROUP'] = df['CATEGORY'].apply(lambda x: 'Income' if x == 'Income' else 'Expenses')
        pivot_df = df.pivot_table(index='DATE', columns='GROUP', values='AMMOUNT', aggfunc='sum', fill_value=0)

        bar_width = 0.4
        bar_pos = np.arange(len(pivot_df.index))
        self.daily_spendings_stat_ax.clear()
        if 'Expenses' in pivot_df.columns:
            expenses = pivot_df['Expenses']
            self.daily_spendings_stat_ax.bar(bar_pos - bar_width/2, expenses, width=bar_width, label='Expenses')
        if 'Income' in pivot_df.columns:
            income = pivot_df['Income']
            self.daily_spendings_stat_ax.bar(bar_pos + bar_width/2, income, width=bar_width, label='Income')

        # Formatting
        self.daily_spendings_stat_ax.set_xticks(range(len(pivot_df.index)))
        self.daily_spendings_stat_ax.set_xticklabels(pivot_df.index, rotation=45, ha='right', fontsize=8)
        self.daily_spendings_stat_ax.legend()

        self.daily_spendings_stat_canvas.draw()

    def get_current_year_data(self, workbook: Workbook):
        data = []
        for sheet in workbook.sheetnames:
            data += self.get_current_month_data(workbook[sheet])
        return data

    def get_current_month_data(self, sheet: Worksheet):
        data = []
        self.headings = [cell.value for cell in sheet[1][:5]]
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=5, values_only=True):
            if row[0] is not None:
                if type(row[0]) is not str: # to ignore the first row
                    row = (row[0].strftime("%d/%m/%Y"), row[1], row[2], row[3], row[4])
                data.append(row)
                self.row_nmb += 1
        return data

    def perform_op(self):
        self.performing_operation = True

    def finish_op(self):
        self.performing_operation = False

    def on_closing(self):
        if not self.performing_operation:
            self.current_workbook.close()
            self.stop()
            self.root.destroy()
            self.print_logs()

    def add_log(self, message: str):
        self.log.append(f'{self.get_timestamp()} - {message}')

    def get_timestamp(self):
        return datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    def print_logs(self):
        for log in self.log: print(log)

    def run(self):
        self.root.mainloop()

    def stop(self):
        self.root.quit()

# some imports down here due to circular dependency
import EditMonth as editMonth
import EditData as editFrame
import AddMonth as addMonth
import AddData as addFrame
import AddYear as addYear

# Manager().run()